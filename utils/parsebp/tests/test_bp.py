import pytest
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import sys
sys.path.insert(0, './')

from parsebp.bosch_price import BoschPrice, UnsupportedFileStructureError
from parsebp import models


class TestBoschPrice:

    regular_file = 'parsebp/tests/regular.xlsx'
    # regular_file = 'parsebp/tests/regular.xlsx'
    lazy_file = 'parsebp/tests/lazy.xlsx'
    # lazy_file = 'parsebp/tests/lazy.xlsx'
    no_master_data_file = 'parsebp/tests/no_master_data.xlsx'
    # no_master_data_file = 'parsebp/tests/no_master_data.xlsx'

    unsupported_file = 'parsebp/tests/' + os.path.basename(__file__)
    unsupported_file_structure = 'parsebp/tests/unsupported_file_structure.xlsx'

    def test_unsupported_load(self):
        with pytest.raises(InvalidFileException):
            BoschPrice(load_workbook(self.unsupported_file, read_only=True))

    @classmethod
    def normal_load(cls, file=regular_file):
        bp = BoschPrice(load_workbook(file, read_only=True))
        assert isinstance(bp.wb, Workbook), f'Result must be openpyxl.Workbook instance. ' \
                                            f'{type(bp)} given.'
        return bp

    @pytest.mark.parametrize(
        'bp_file, expected',
        [
            (
                regular_file,
                [
                    ('18.04.2022_Pricelist_UA', models.PriceList),
                    ('NEW RELEASE', models.NewRelease),
                    ('Master data', models.MasterData),
                    ('Зняті з поставок', models.Discontinued),
                    ('Заміни', models.References),
                ]
            ),
            (
                lazy_file,
                [
                    ('01.07.2022_Pricelist_UA', models.PriceList),
                    ('Master data', models.MasterData),
                ]
            ),
            (
                no_master_data_file,
                [
                    ('01.09.2022_Pricelist_UA', models.PriceList),
                ]
            ),
        ]
    )
    def test_supported_file_structure(self, bp_file, expected):
        bp = self.normal_load(bp_file)
        assert set(bp.mapped_sheets()) == set(expected)

    def test_unsupported_file_structure(self):
        bp = self.normal_load(self.unsupported_file_structure)
        with pytest.raises(UnsupportedFileStructureError):
            bp.mapped_sheets()
