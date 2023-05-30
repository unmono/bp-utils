import re
import os
import sqlite3
from typing import Tuple, Protocol, ClassVar
from contextlib import closing

import pydantic.error_wrappers
from openpyxl import Workbook

from parsebp.models import (
    PriceList,
    MasterData,
    NewRelease,
    Discontinued,
    References,
)

from settings import ParseSettings
settings = ParseSettings()


class UnsupportedFileStructureError(Exception):
    pass


class BPModelTYpe(Protocol):
    sheet_pattern: ClassVar[re.Pattern]


class BoschPrice:
    """
    Represents a loaded Bosch price list.
    """

    def __init__(self, wb: Workbook):
        self.wb = wb
        self.sheet_names = wb.sheetnames

        self.required_model = PriceList
        self.extra_models = [
            MasterData,
            NewRelease,
            Discontinued,
            References,
        ]

    def map_model_to_sheet(self, model: BPModelTYpe) -> Tuple[str, BPModelTYpe] | None:
        """
        Searches in sheet names of the loaded file for patterns of models
        :param model: bp model
        :return: Tuple('Corresponding sheet name', bp model)
        """
        for sheet in self.sheet_names:
            if model.sheet_pattern.search(sheet):
                return sheet, model
        return None

    def mapped_sheets(self):
        if (required_sheet := self.map_model_to_sheet(self.required_model)) is None:
            raise UnsupportedFileStructureError('Unsupported file structure')

        extra_sheets = [
            t for model in self.extra_models
            if (t := self.map_model_to_sheet(model)) is not None
        ]

        return [required_sheet, ] + extra_sheets

    @staticmethod
    def prepare_db(mapped_sheets) -> None:
        try:
            os.remove(settings.RAW_DB)
        except FileNotFoundError:
            pass

        with closing(sqlite3.connect(settings.RAW_DB)) as db:
            for _, model in mapped_sheets:
                db.execute(model.sqlite_schema())

    def populate_db(self):
        ms = self.mapped_sheets()
        self.prepare_db(ms)

        for sheet, model in ms:
            ws = self.wb[sheet]
            set_of_values = []
            schema = model.schema()['properties'].items()
            for i in range(2, ws.max_row + 1):
                data = {
                    t: ws[p['excel_column'] + str(i)].value
                    for t, p in schema
                }
                try:
                    obj = model(**data)
                except pydantic.error_wrappers.ValidationError as e:
                    print(i)
                    raise e

                set_of_values.append(tuple(obj.dict().values()))

            title = model.sql_title()
            placeholders = ', '.join(['?'] * len(set_of_values[0]))

            with closing(sqlite3.connect(settings.RAW_DB)) as db:
                with db:
                    db.executemany(f"""
                        INSERT INTO {title} 
                        VALUES({placeholders})
                    """, set_of_values)
